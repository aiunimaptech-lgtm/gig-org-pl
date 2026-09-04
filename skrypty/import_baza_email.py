"""
Import bazy e-mail do Supabase (tabela baza_email) z pliku xlsx.

Użycie:
  set GIG_IMPORT_TOKEN=<token z private.gig_sekrety 'import_token'>
  python skrypty/import_baza_email.py "sciezka\\do\\dane_firm_z_maili_GIG.xlsx"

Czyta arkusz „Dane firm" (kolumny: E-mail, Pochodzenie, Grupa, rodzaj/branża,
Firma / instytucja, Adres, NIP, Telefon, Osoba kontaktowa, Stanowisko osoby
kontaktowej, WWW) i dokłada „Pewność" + „Źródło danych" z arkusza
„Do weryfikacji" (po e-mailu). Wysyła paczkami do RPC gig_baza_email_import
(upsert po e-mailu — puste pola nie nadpisują istniejących).
Token jest jednorazowy: backend/supabase_baza_email.sql.
Wymaga: pip install openpyxl requests
"""
import json, os, sys, warnings
import openpyxl, requests

SUPABASE_URL = "https://zlepwzeyjwpmhyxfnime.supabase.co"
ANON = "sb_publishable_1KPF4mdln3C-cZHMIqAOFw_ZM8Go2Ji"   # klucz publiczny (jak w gig-config.js)
PACZKA = 400

KOLUMNY = {
    "E-mail": "email", "Pochodzenie": "pochodzenie", "Grupa": "grupa", "rodzaj/branża": "rodzaj",
    "Firma / instytucja": "firma", "Adres": "adres", "NIP": "nip", "Telefon": "telefon",
    "Osoba kontaktowa": "osoba", "Stanowisko osoby kontaktowej": "stanowisko", "WWW": "www",
    "Pewność": "pewnosc", "Źródło danych": "zrodlo",
}

def s(v):
    if v is None: return None
    v = str(v).strip()
    return v or None

def wczytaj(sciezka):
    warnings.simplefilter("ignore")
    wb = openpyxl.load_workbook(sciezka, read_only=True, data_only=True)
    def arkusz(nazwa):
        if nazwa not in wb.sheetnames: return []
        rows = list(wb[nazwa].iter_rows(values_only=True))
        if not rows: return []
        hdr = [KOLUMNY.get(str(h).strip(), None) if h else None for h in rows[0]]
        out = []
        for r in rows[1:]:
            if not any(r): continue
            rec = {k: s(v) for k, v in zip(hdr, r) if k}
            if rec.get("email"): out.append(rec)
        return out
    dane = arkusz("Dane firm")
    wer = {r["email"].lower(): r for r in arkusz("Do weryfikacji")}
    for r in dane:
        w = wer.get(r["email"].lower())
        if w:
            for k in ("pewnosc", "zrodlo"):
                if w.get(k) and not r.get(k): r[k] = w[k]
    return dane

def main():
    if len(sys.argv) < 2: sys.exit("podaj sciezke do xlsx")
    token = os.environ.get("GIG_IMPORT_TOKEN", "").strip()
    if len(token) != 64: sys.exit("brak GIG_IMPORT_TOKEN (64 znaki hex)")
    rekordy = wczytaj(sys.argv[1])
    print(f"rekordow do importu: {len(rekordy)}")
    razem = 0
    for i in range(0, len(rekordy), PACZKA):
        paczka = rekordy[i:i + PACZKA]
        res = requests.post(f"{SUPABASE_URL}/rest/v1/rpc/gig_baza_email_import",
                            headers={"apikey": ANON, "Authorization": f"Bearer {ANON}", "Content-Type": "application/json"},
                            data=json.dumps({"p_token": token, "p_rows": paczka}, ensure_ascii=False).encode("utf-8"),
                            timeout=120)
        if res.status_code != 200:
            sys.exit(f"paczka {i//PACZKA + 1}: HTTP {res.status_code}: {res.text[:300]}")
        n = res.json(); razem += int(n)
        print(f"  paczka {i//PACZKA + 1}: {len(paczka)} -> zapisano {n}")
    print(f"GOTOWE: {razem} wierszy")

if __name__ == "__main__":
    main()
